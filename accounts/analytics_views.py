from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.db.models import Sum, Count, F
from django.db.models.functions import TruncDay, TruncWeek, TruncMonth, TruncQuarter, TruncYear
from django.utils import timezone
from booking.models import Booking
from accounts.models import User
from tenants.utils import has_tenant_permission
from django.shortcuts import redirect
import datetime
from services.models import GuestOrder
from events.models import EventBooking
from gym.models import GymMembership
from core.models import TenantSetting
import os
import tempfile

import json
from django.core.serializers.json import DjangoJSONEncoder

class HotelStatisticsView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'analytics/hotel_statistics.html'

    def test_func(self):
        # Allow Admin and Manager
        tenant = getattr(self.request, 'tenant', None)
        if not tenant: return False
        return has_tenant_permission(self.request.user, tenant, ['ADMIN', 'MANAGER'])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tenant = self.request.tenant
        now = timezone.now()
        
        # Base QuerySets
        bookings = Booking.objects.filter(tenant=tenant)
        orders = GuestOrder.objects.filter(booking__tenant=tenant).exclude(status='CANCELLED')
        events = EventBooking.objects.filter(hall__tenant=tenant).exclude(status='CANCELLED')
        gym_memberships = GymMembership.objects.filter(plan__tenant=tenant).exclude(status='CANCELLED')
        
        # 1. Overview Cards (Total Lifetime)
        # Note: Ideally we should use Payment model for cash flow, but aggregated model sums work for accrued revenue
        
        room_revenue = bookings.aggregate(total=Sum('total_price'))['total'] or 0
        service_revenue = orders.aggregate(total=Sum('total_price'))['total'] or 0
        event_revenue = events.aggregate(total=Sum('total_price'))['total'] or 0
        
        # Gym revenue calculation (price is on Plan)
        # We need to sum plan.price for all memberships
        gym_revenue = 0
        for m in gym_memberships:
            if m.plan:
                gym_revenue += m.plan.price
        
        total_revenue = room_revenue + service_revenue + event_revenue + gym_revenue
        
        context['total_sales'] = total_revenue
        context['total_guests'] = bookings.values('guest_email').distinct().count()
        context['total_bookings'] = bookings.count()
        
        # Revenue Breakdown for Charts
        revenue_breakdown = {
            'Rooms': float(room_revenue),
            'Services': float(service_revenue),
            'Events': float(event_revenue),
            'Gym': float(gym_revenue)
        }
        context['revenue_breakdown'] = revenue_breakdown
        context['revenue_breakdown_json'] = json.dumps(revenue_breakdown, cls=DjangoJSONEncoder)
        
        # 2. Aggregations (Sales & Guests)
        # We will focus on Booking trends for the main graphs as combining all temporal data is complex for this scope
        # unless we use a unified Transaction model.
        
        # Helper to get data
        def get_aggregated_data(trunc_func, date_field='created_at'):
            return bookings.annotate(
                period=trunc_func(date_field)
            ).values('period').annotate(
                sales=Sum('total_price'),
                guests=Count('id') 
            ).order_by('period') # Chronological order for charts

        # Per Day (Last 30 Days)
        daily_stats = list(get_aggregated_data(TruncDay))[-30:]
        context['daily_stats'] = daily_stats[::-1] # Reverse for table (Newest first)
        
        # Per Week (Last 12 Weeks)
        weekly_stats = list(get_aggregated_data(TruncWeek))[-12:]
        context['weekly_stats'] = weekly_stats[::-1]
        
        # Per Month (Last 12 Months)
        monthly_stats = list(get_aggregated_data(TruncMonth))[-12:]
        context['monthly_stats'] = monthly_stats[::-1]
        
        # Per Quarter
        quarterly_stats = list(get_aggregated_data(TruncQuarter))[-8:]
        context['quarterly_stats'] = quarterly_stats[::-1]
        
        # Per Year
        yearly_stats = list(get_aggregated_data(TruncYear))
        context['yearly_stats'] = yearly_stats[::-1]
        
        # Filter Logic
        period = self.request.GET.get('period', 'all')
        context['selected_period'] = period
        
        # Prepare Chart Data
        # 1. Revenue Breakdown
        context['breakdown_labels'] = json.dumps(['Room Bookings', 'Room Service', 'Events', 'Gym'])
        context['breakdown_values'] = json.dumps([
            float(room_revenue), float(service_revenue), float(event_revenue), float(gym_revenue)
        ])
        
        # 2. Sales Trend
        # Select data based on filter, default to daily (last 30 days) if 'all' or 'daily'
        trend_data = context['daily_stats']
        trend_label_fmt = '%Y-%m-%d'
        
        if period == 'weekly': 
            trend_data = context['weekly_stats']
            trend_label_fmt = 'Week %W'
        elif period == 'monthly': 
            trend_data = context['monthly_stats']
            trend_label_fmt = '%b %Y'
        elif period == 'quarterly': 
            trend_data = context['quarterly_stats']
            trend_label_fmt = 'Q%q %Y' # Custom handling needed for JS usually, but simple strftime fallback
        elif period == 'yearly': 
            trend_data = context['yearly_stats']
            trend_label_fmt = '%Y'
            
        # Prepare lists (reverse to show chronological order left-to-right)
        chart_labels = []
        chart_values = []
        
        for item in reversed(trend_data):
            # Format label
            d = item['period']
            if period == 'quarterly':
                q = (d.month - 1) // 3 + 1
                lbl = f"Q{q} {d.year}"
            elif period == 'weekly':
                lbl = f"W{d.strftime('%W')} {d.year}"
            else:
                lbl = d.strftime(trend_label_fmt)
            
            chart_labels.append(lbl)
            chart_values.append(float(item['sales'] or 0))
            
        context['chart_labels'] = json.dumps(chart_labels)
        context['chart_values'] = json.dumps(chart_values)
        
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            tables_html = render_to_string('analytics/partials/tables.html', context, request=self.request)
            return JsonResponse({
                'chart_labels': chart_labels,
                'chart_values': chart_values,
                'tables_html': tables_html
            })
        
        return context

from django.http import HttpResponse
from fpdf import FPDF
import io
import csv

def download_statistics_report(request):
    # Permission Check
    if not request.user.is_authenticated:
        return redirect('login')
    
    tenant = getattr(request, 'tenant', None)
    if not tenant:
        return redirect('home')
        
    if not has_tenant_permission(request.user, tenant, ['ADMIN', 'MANAGER']):
        return redirect('dashboard')

    # Get Filter
    period_filter = request.GET.get('period', 'all')

    # Gather Data
    bookings = Booking.objects.filter(tenant=tenant)
    orders = GuestOrder.objects.filter(booking__tenant=tenant).exclude(status='CANCELLED')
    events = EventBooking.objects.filter(hall__tenant=tenant).exclude(status='CANCELLED')
    gym_memberships = GymMembership.objects.filter(plan__tenant=tenant).exclude(status='CANCELLED')
    
    # Financials
    room_revenue = float(bookings.aggregate(total=Sum('total_price'))['total'] or 0)
    service_revenue = float(orders.aggregate(total=Sum('total_price'))['total'] or 0)
    event_revenue = float(events.aggregate(total=Sum('total_price'))['total'] or 0)
    gym_revenue = float(sum(m.plan.price for m in gym_memberships if m.plan))
    total_sales = room_revenue + service_revenue + event_revenue + gym_revenue
    total_bookings = bookings.count()
    
    # Helper
    def get_data(trunc_func):
        return bookings.annotate(
            period=trunc_func('created_at')
        ).values('period').annotate(
            sales=Sum('total_price'),
            guests=Count('id')
        ).order_by('-period')

    daily = get_data(TruncDay)[:30]
    weekly = get_data(TruncWeek)[:12]
    monthly = get_data(TruncMonth)[:12]
    quarterly = get_data(TruncQuarter)[:4]
    yearly = get_data(TruncYear)[:5]
    
    # PDF Generation
    # Get Theme Settings
    settings = TenantSetting.objects.filter(tenant=tenant).first()
    current_theme = settings.theme if settings else 'theme-default'
    hotel_name = settings.hotel_name if settings else tenant.name
    
    # Define Theme Colors (R, G, B)
    THEME_COLORS = {
        'theme-default': (19, 236, 109),
        'theme-light': (19, 236, 109), 
        'theme-blue': (59, 130, 246), 
        'theme-luxury': (212, 175, 55),
        'theme-forest': (74, 222, 128),
        'theme-ocean': (6, 182, 212),
        'theme-sunset': (244, 114, 182),
        'theme-royal': (167, 139, 250),
        'theme-minimal': (113, 113, 122),
    }
    primary_color = THEME_COLORS.get(current_theme, (19, 236, 109))

    pdf = FPDF()
    pdf.add_page()
    
    # --- Decorative Header (Matching Receipt) ---
    pdf.set_y(10)
    pdf.set_font("Arial", 'B', 16)
    pdf.set_text_color(*primary_color)
    pdf.cell(0, 8, txt=hotel_name.upper(), ln=1, align="R")
    
    pdf.set_font("Arial", '', 8)
    pdf.set_text_color(60, 60, 60)
    pdf.cell(0, 4, txt="Statistics Report", ln=1, align="R")
    
    pdf.ln(2)
    pdf.set_draw_color(*primary_color)
    pdf.set_line_width(0.5)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y()) # A4 width approx 210mm
    pdf.set_line_width(0.2)
    
    pdf.ln(10)
    
    # Report Title & Date
    pdf.set_font("Arial", 'B', 20)
    pdf.set_text_color(30, 41, 59)
    pdf.cell(120, 10, txt=f"{period_filter.title()} Report" if period_filter != 'all' else "Full Statistics Report", ln=0, align="L")
    
    pdf.set_fill_color(240, 240, 240)
    pdf.set_font("Arial", 'B', 10)
    pdf.cell(70, 10, txt=f"Generated: {timezone.now().strftime('%Y-%m-%d')}", ln=1, align="R", fill=True)
    pdf.ln(10)
    
    # --- Helper: Draw PDF Line Chart ---
    def draw_line_chart(pdf, title, labels, values, x_pos, y_pos, w, h):
        if not values or len(values) < 2: return
        
        pdf.set_xy(x_pos, y_pos)
        pdf.set_font("Arial", 'B', 10)
        pdf.set_text_color(30, 41, 59)
        pdf.cell(w, 8, title, ln=1, align='C')
        
        # Chart Area
        chart_y = y_pos + 12
        chart_h = h - 20
        chart_w = w
        max_val = max(values) * 1.1 if max(values) > 0 else 1
        
        # Axis
        pdf.set_draw_color(200, 200, 200)
        pdf.line(x_pos, chart_y + chart_h, x_pos + chart_w, chart_y + chart_h) # X-axis
        pdf.line(x_pos, chart_y, x_pos, chart_y + chart_h) # Y-axis
        
        # Plot Lines
        pdf.set_draw_color(*primary_color)
        pdf.set_line_width(0.8)
        
        step_w = chart_w / (len(values) - 1)
        
        prev_x, prev_y = None, None
        
        for i, val in enumerate(values):
            # Calculate positions
            px = x_pos + (i * step_w)
            py = chart_y + chart_h - ((val / max_val) * chart_h)
            
            # Draw Line
            if prev_x is not None:
                pdf.line(prev_x, prev_y, px, py)
                
            # Draw Point
            pdf.set_fill_color(*primary_color)
            pdf.circle(px, py, 1, 'F') # Requires newer FPDF or custom circle. 
            # If circle not available in basic FPDF, use rect
            # pdf.rect(px-1, py-1, 2, 2, 'F') 
            
            # Label (Sparse if too many)
            if len(values) > 15:
                if i % 3 == 0:
                    pdf.set_font("Arial", '', 6)
                    pdf.set_xy(px - 3, chart_y + chart_h + 1)
                    pdf.cell(6, 4, labels[i], 0, 0, 'C')
            else:
                pdf.set_font("Arial", '', 6)
                pdf.set_xy(px - 3, chart_y + chart_h + 1)
                pdf.cell(6, 4, labels[i], 0, 0, 'C')
                
            prev_x, prev_y = px, py
            
    # --- Helper: Draw Breakdown Chart (Side-by-side Bars) ---
    def draw_breakdown_chart(pdf, data, x_pos, y_pos, w, h):
        pdf.set_xy(x_pos, y_pos)
        pdf.set_font("Arial", 'B', 10)
        pdf.set_text_color(30, 41, 59)
        pdf.cell(w, 8, "Revenue Breakdown", ln=1, align='C')
        
        y = y_pos + 12
        max_w = w - 30
        total = sum(data.values())
        if total == 0: total = 1
        
        colors = [
            (19, 236, 109), # Green
            (59, 130, 246), # Blue
            (245, 158, 11), # Orange
            (236, 72, 153)  # Pink
        ]
        
        i = 0
        for label, val in data.items():
            pct = val / total
            bar_w = pct * max_w
            
            # Color Box
            color = colors[i % len(colors)]
            pdf.set_fill_color(*color)
            pdf.rect(x_pos, y + 1, 4, 4, 'F')
            
            # Label & Value
            pdf.set_xy(x_pos + 6, y)
            pdf.set_font("Arial", '', 8)
            pdf.cell(30, 6, f"{label} ({pct*100:.1f}%)", 0, 0)
            
            # Bar (Visual)
            # pdf.rect(x_pos + 40, y + 1, bar_w, 4, 'F')
            
            y += 8
            i += 1
            
        # Draw Pie-ish visualization using arcs if possible or just colored bars stack
        # For simplicity in FPDF 1.7: Stacked Bar
        y += 5
        start_x = x_pos
        full_w = w - 10
        pdf.set_xy(start_x, y)
        
        i = 0
        for label, val in data.items():
            pct = val / total
            if pct > 0:
                seg_w = pct * full_w
                color = colors[i % len(colors)]
                pdf.set_fill_color(*color)
                pdf.rect(start_x, y, seg_w, 8, 'F')
                start_x += seg_w
            i += 1

    # Draw Charts Area
    # Two Columns: Trend (Left), Breakdown (Right)
    
    chart_area_y = pdf.get_y() + 5
    col_width = 90
    chart_height = 60
    
    # Trend Data Preparation
    chart_labels = []
    chart_data = []
    chart_title = "Sales Trend"
    
    if period_filter in ['all', 'daily']:
        stats = list(daily)
        stats.reverse() 
        chart_title = "Daily Sales Trend"
        for x in stats:
            chart_labels.append(x['period'].strftime('%d')) 
            chart_data.append(float(x['sales'] or 0))
    elif period_filter == 'weekly':
        stats = list(weekly)
        stats.reverse()
        chart_title = "Weekly Sales Trend"
        for x in stats:
            chart_labels.append(x['period'].strftime('W%W'))
            chart_data.append(float(x['sales'] or 0))
    elif period_filter == 'monthly':
        stats = list(monthly)
        stats.reverse()
        chart_title = "Monthly Sales Trend"
        for x in stats:
            chart_labels.append(x['period'].strftime('%b'))
            chart_data.append(float(x['sales'] or 0))
            
    # Draw Line Chart (Left)
    draw_line_chart(pdf, chart_title, chart_labels, chart_data, 10, chart_area_y, col_width, chart_height)
    
    # Draw Breakdown Chart (Right)
    breakdown_data = {
        'Rooms': room_revenue,
        'Service': service_revenue,
        'Events': event_revenue,
        'Gym': gym_revenue
    }
    draw_breakdown_chart(pdf, breakdown_data, 110, chart_area_y, col_width, chart_height)
    
    pdf.set_y(chart_area_y + chart_height + 10)

    # Overview Section (Only show if 'all' or specific requests? Let's show it always as summary)
    pdf.set_fill_color(*primary_color)

    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Arial", 'B', 10)
    pdf.cell(0, 8, "  FINANCIAL OVERVIEW", ln=True, fill=True)
    
    pdf.set_text_color(30, 41, 59)
    pdf.set_font("Arial", '', 10)
    pdf.ln(2)
    
    def print_row(label, value):
        pdf.cell(100, 8, label, 0, 0, 'L')
        pdf.cell(90, 8, value, 0, 1, 'R')
        pdf.set_draw_color(230, 230, 230)
        pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    
    print_row("Total Revenue", f"{total_sales:,.2f}")
    print_row(" - Room Bookings", f"{room_revenue:,.2f}")
    print_row(" - Room Service", f"{service_revenue:,.2f}")
    print_row(" - Events", f"{event_revenue:,.2f}")
    print_row(" - Gym Memberships", f"{gym_revenue:,.2f}")
    pdf.ln(2)
    print_row("Total Room Bookings", str(total_bookings))
    pdf.ln(10)
    
    # Helper for Tables
    def add_table(title, data, date_fmt):
        if not data: return
        
        pdf.set_fill_color(30, 41, 59)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Arial", 'B', 10)
        pdf.cell(0, 8, f"  {title.upper()}", ln=True, fill=True)
        pdf.ln(2)
        
        # Header
        pdf.set_font("Arial", 'B', 9)
        pdf.set_fill_color(240, 240, 240)
        pdf.set_text_color(30, 41, 59)
        
        pdf.cell(80, 8, "Period", 0, 0, 'L', 1)
        pdf.cell(55, 8, "Sales", 0, 0, 'R', 1)
        pdf.cell(55, 8, "Guests/Bookings", 0, 1, 'R', 1)
        
        # Rows
        pdf.set_font("Arial", '', 9)
        fill = False
        for row in data:
            if "%q" in date_fmt:
                q = (row['period'].month - 1) // 3 + 1
                date_str = f"Q{q} {row['period'].year}"
            else:
                date_str = row['period'].strftime(date_fmt)
                
            sales_val = row['sales'] or 0
            sales = f"{sales_val:,.2f}"
            guests = str(row['guests'])
            
            # Alternating rows logic if desired, but clean white is fine too
            pdf.cell(80, 8, date_str, 0, 0, 'L')
            pdf.cell(55, 8, sales, 0, 0, 'R')
            pdf.cell(55, 8, guests, 0, 1, 'R')
            
            pdf.set_draw_color(230, 230, 230)
            pdf.line(10, pdf.get_y(), 200, pdf.get_y())
            
        pdf.ln(8)

    # Filtered Content
    if period_filter in ['all', 'daily']:
        add_table("Daily Room Stats (Last 30 Days)", daily, "%Y-%m-%d")
        
    if period_filter in ['all', 'weekly']:
        add_table("Weekly Room Stats (Last 12 Weeks)", weekly, "Week %W, %Y")
        
    if period_filter in ['all', 'monthly']:
        add_table("Monthly Room Stats (Last 12 Months)", monthly, "%B %Y")
        
    if period_filter in ['all', 'quarterly']:
        add_table("Quarterly Room Stats", quarterly, "%q")
        
    if period_filter in ['all', 'yearly']:
        # Yearly Table
        pdf.set_fill_color(30, 41, 59)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Arial", 'B', 10)
        pdf.cell(0, 8, "  YEARLY ROOM STATS", ln=True, fill=True)
        pdf.ln(2)
        
        pdf.set_font("Arial", 'B', 9)
        pdf.set_fill_color(240, 240, 240)
        pdf.set_text_color(30, 41, 59)
        pdf.cell(80, 8, "Year", 0, 0, 'L', 1)
        pdf.cell(55, 8, "Sales", 0, 0, 'R', 1)
        pdf.cell(55, 8, "Guests", 0, 1, 'R', 1)
        
        pdf.set_font("Arial", '', 9)
        for row in yearly:
            pdf.cell(80, 8, row['period'].strftime("%Y"), 0, 0, 'L')
            sales_val = row['sales'] or 0
            pdf.cell(55, 8, f"{sales_val:,.2f}", 0, 0, 'R')
            pdf.cell(55, 8, str(row['guests']), 0, 1, 'R')
            pdf.set_draw_color(230, 230, 230)
            pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    
    # Output
    # FIX: Remove .encode('latin-1') because pdf.output(dest='S') returns bytearray in newer FPDF versions
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_pdf:
        tmp_pdf_path = tmp_pdf.name

    try:
        pdf.output(name=tmp_pdf_path, dest='F')
        with open(tmp_pdf_path, 'rb') as f:
            pdf_content = f.read()
            
        response = HttpResponse(pdf_content, content_type='application/pdf')
        filename = f"Statistics_{period_filter}_{timezone.now().date()}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    finally:
        if os.path.exists(tmp_pdf_path):
            os.unlink(tmp_pdf_path)

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference, Series

def download_statistics_excel(request):
    # Permission Check
    if not request.user.is_authenticated:
        return redirect('login')
    
    tenant = getattr(request, 'tenant', None)
    if not tenant:
        return redirect('home')
        
    if not has_tenant_permission(request.user, tenant, ['ADMIN', 'MANAGER']):
        return redirect('dashboard')
        
    # Get Filter
    period_filter = request.GET.get('period', 'all')

    # Gather Data
    bookings = Booking.objects.filter(tenant=tenant)
    orders = GuestOrder.objects.filter(booking__tenant=tenant).exclude(status='CANCELLED')
    events = EventBooking.objects.filter(hall__tenant=tenant).exclude(status='CANCELLED')
    gym_memberships = GymMembership.objects.filter(plan__tenant=tenant).exclude(status='CANCELLED')
    
    room_revenue = bookings.aggregate(total=Sum('total_price'))['total'] or 0
    service_revenue = orders.aggregate(total=Sum('total_price'))['total'] or 0
    event_revenue = events.aggregate(total=Sum('total_price'))['total'] or 0
    gym_revenue = sum(m.plan.price for m in gym_memberships if m.plan)
    total_sales = room_revenue + service_revenue + event_revenue + gym_revenue
    
    # Create Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Statistics Report"
    
    # --- Styles ---
    header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    subheader_font = Font(name='Arial', size=12, bold=True, color='333333')
    text_font = Font(name='Arial', size=10)
    bold_text_font = Font(name='Arial', size=10, bold=True)
    
    header_fill = PatternFill(start_color='13EC6D', end_color='13EC6D', fill_type='solid') # Primary Green
    subheader_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid') # Light Grey
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    # --- Report Header ---
    ws.merge_cells('A1:C1')
    cell = ws['A1']
    cell.value = f"Hotel Statistics Report - {tenant.name}"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    
    ws.merge_cells('A2:C2')
    cell = ws['A2']
    cell.value = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
    cell.font = Font(name='Arial', size=10, italic=True)
    cell.alignment = center_align
    
    # --- Financial Overview ---
    row = 4
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws[f'A{row}']
    cell.value = "Financial Overview"
    cell.font = subheader_font
    cell.fill = subheader_fill
    cell.alignment = left_align
    row += 1
    
    # Table Header
    headers = ['Category', 'Revenue', 'Percentage']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = bold_text_font
        cell.border = thin_border
        cell.alignment = center_align
    row += 1
    
    # Data Rows
    data = [
        ('Room Bookings', room_revenue),
        ('Room Service', service_revenue),
        ('Events', event_revenue),
        ('Gym Memberships', gym_revenue)
    ]
    
    total_val = float(total_sales) if total_sales > 0 else 1
    
    for category, value in data:
        val = float(value)
        pct = (val / total_val)
        
        c1 = ws.cell(row=row, column=1, value=category)
        c1.border = thin_border
        c1.alignment = left_align
        
        c2 = ws.cell(row=row, column=2, value=val)
        c2.number_format = '#,##0.00'
        c2.border = thin_border
        
        c3 = ws.cell(row=row, column=3, value=pct)
        c3.number_format = '0.0%'
        c3.border = thin_border
        
        row += 1
        
    # Total Row
    ws.cell(row=row, column=1, value="TOTAL").font = bold_text_font
    ws.cell(row=row, column=1).border = thin_border
    
    c_total = ws.cell(row=row, column=2, value=total_sales)
    c_total.font = bold_text_font
    c_total.number_format = '#,##0.00'
    c_total.border = thin_border
    
    ws.cell(row=row, column=3, value="100%").border = thin_border
    
    # --- Pie Chart (Revenue Breakdown) ---
    pie = PieChart()
    pie.title = "Revenue Breakdown"
    labels = Reference(ws, min_col=1, min_row=6, max_row=9)
    data = Reference(ws, min_col=2, min_row=6, max_row=9)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.height = 7.5 # cm approx
    pie.width = 12 # cm approx
    
    # Chart Grid Logic
    # Charts start at Column F (Index 6)
    # Grid: 3 charts per row
    # Cell Width approx: A=30, B=20, C=20 -> Total 70. 
    # E is buffer. F starts charts.
    
    chart_counter = 0
    
    # Position Pie Chart (First Chart)
    # Row 1, Col 1 of Grid
    # Anchor: F4
    ws.add_chart(pie, "F4")
    chart_counter += 1

    row += 3
    
    # --- Helper for Data Tables ---
    def add_excel_table(title, qs, period_trunc, date_fmt):
        nonlocal row, chart_counter
        start_row = row
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = title
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = left_align
        row += 1
        
        headers = ['Period', 'Sales', 'Guests/Bookings']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = bold_text_font
            cell.border = thin_border
            cell.alignment = center_align
        row += 1
        
        stats = bookings.annotate(
            period=period_trunc('created_at')
        ).values('period').annotate(
            sales=Sum('total_price'),
            guests=Count('id')
        ).order_by('-period')
        
        if 'Day' in str(period_trunc): stats = stats[:30]
        elif 'Week' in str(period_trunc): stats = stats[:12]
        elif 'Month' in str(period_trunc): stats = stats[:12]
        elif 'Quarter' in str(period_trunc): stats = stats[:4]
        elif 'Year' in str(period_trunc): stats = stats[:5]
        
        data_start_row = row
        count = 0
        
        for item in stats:
            count += 1
            if "%q" in date_fmt:
                 q = (item['period'].month - 1) // 3 + 1
                 date_str = f"Q{q} {item['period'].year}"
            else:
                 date_str = item['period'].strftime(date_fmt)
                 
            sales = float(item['sales'] or 0)
            guests = item['guests']
            
            c1 = ws.cell(row=row, column=1, value=date_str)
            c1.alignment = center_align
            c1.border = thin_border
            
            c2 = ws.cell(row=row, column=2, value=sales)
            c2.number_format = '#,##0.00'
            c2.border = thin_border
            
            c3 = ws.cell(row=row, column=3, value=guests)
            c3.alignment = center_align
            c3.border = thin_border
            
            row += 1
            
        data_end_row = row - 1
        
        if count > 0:
            # --- Bar Chart (Sales Trend) ---
            chart = BarChart()
            chart.title = title.replace("Stats", "Trend")
            chart.style = 10
            chart.x_axis.title = 'Period'
            chart.y_axis.title = 'Sales'
            chart.height = 7.5 
            chart.width = 12 
            
            cats = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
            data = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_end_row)
            
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(cats)
            
            # Position Chart in Grid
            # Grid starts at F4.
            # 3 Columns wide (F, L, R approx)
            # Rows spacing approx 16 rows.
            
            # Calculate Grid Position
            grid_row = chart_counter // 3
            grid_col = chart_counter % 3
            
            # Base Row = 4. Height spacing = 16 rows.
            anchor_row = 4 + (grid_row * 16)
            
            # Base Col = F (Index 6). Width spacing = 8 columns (approx width of chart in cells)
            anchor_col_idx = 6 + (grid_col * 8)
            anchor_col_letter = get_column_letter(anchor_col_idx)
            
            ws.add_chart(chart, f"{anchor_col_letter}{anchor_row}")
            chart_counter += 1

        row += 2

    # --- Daily Stats ---
    if period_filter in ['all', 'daily']:
        add_excel_table("Daily Room Stats (Last 30 Days)", bookings, TruncDay, "%Y-%m-%d")
    
    if period_filter in ['all', 'weekly']:
        add_excel_table("Weekly Room Stats (Last 12 Weeks)", bookings, TruncWeek, "Week %W, %Y")
        
    if period_filter in ['all', 'monthly']:
        add_excel_table("Monthly Room Stats", bookings, TruncMonth, "%B %Y")
        
    if period_filter in ['all', 'quarterly']:
        add_excel_table("Quarterly Room Stats", bookings, TruncQuarter, "%q")
        
    if period_filter in ['all', 'yearly']:
        add_excel_table("Yearly Room Stats", bookings, TruncYear, "%Y")

    # Adjust Column Widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    
    # Output
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Statistics_{period_filter}_{timezone.now().date()}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
